import datetime as dt
import time
from io import BytesIO
from pathlib import Path
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import xlsxwriter  # type: ignore

# ------------------------------------------------------------------
# ————— Autenticación con roles —————
if "logged" not in st.session_state:
    st.session_state.logged = False

if not st.session_state.logged:
    st.title("🔒 Iniciar sesión")
    user_input = st.text_input("Usuario")
    pwd_input  = st.text_input("Contraseña", type="password")
    if st.button("Entrar"):
        creds = st.secrets["credentials"]
        match = next((info for info in creds.values() if info["username"] == user_input), None)

        if match and pwd_input == match["password"]:
            st.session_state.logged = True
            st.session_state.user   = match["username"]
            st.session_state.role   = match["role"]
            # Rerun automático con la misma sesión
            st.rerun()
        else:
            st.error("Usuario o contraseña inválida")

    # Mientras no haya validado, detener todo lo demás
    st.stop()



# ------------------------------------------------------------------
# 2) Configuración global y helpers
# ------------------------------------------------------------------
EXCEL    = 'Recursos.xlsx'
SHEET    = 'Reservas'
DATE_FMT = '%d/%m/%Y'
BURGUNDY = '#800000'
BLOQUES  = [
    (dt.time(8, 0),  dt.time(9, 30)),
    (dt.time(9, 45), dt.time(11, 15)),
    (dt.time(11, 30), dt.time(13, 0)),
    (dt.time(14, 0),  dt.time(15, 30)),
    (dt.time(15, 45), dt.time(16, 30)),
    (dt.time(16, 30), dt.time(18, 30)),
]

from datetime import date, datetime

def parse_date(val):
    # si ya es date (pero no datetime), lo devolvemos
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    # si es datetime o pandas.Timestamp, devolvemos sólo la fecha
    if isinstance(val, datetime):
        return val.date()

    # si viene como string...
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return None
        # añadir aquí el nuevo patrón:
        allowed_str_formats = [
            "%d/%m/%Y",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",  # <-- lo incorporamos
        ]
        for fmt in allowed_str_formats:
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue

    raise ValueError(f"Formato de fecha inválido: {val!r}")

def as_time(val):
    if isinstance(val, dt.time):
        return val
    for fmt in ('%H:%M:%S', '%H:%M'):
        try:
            return dt.datetime.strptime(str(val), fmt).time()
        except ValueError:
            continue
    raise ValueError(f'Formato de hora inválido: {val}')

def overlap(hi1, hf1, hi2, hf2):
    return max(hi1, hi2) < min(hf1, hf2)

# ------------------------------------------------------------------
# 3) Recarga de datos inicial y listas dinámicas
# ------------------------------------------------------------------
# Inicializar Excel si no existe
df = None
if not Path(EXCEL).exists():
    cols = ['Fecha', 'Hora inicio', 'Hora fin', 'Profesor', 'Curso', 'Recurso', 'Observaciones']
    pd.DataFrame(columns=cols).to_excel(EXCEL, sheet_name=SHEET, index=False)
# Cargar datos
xl = pd.ExcelFile(EXCEL)
df = pd.read_excel(xl, sheet_name=SHEET, dtype=str).fillna('')

def recalc_lists(df: pd.DataFrame):
    get_col = lambda s: (
        xl.parse(s).iloc[:, 0].dropna().astype(str).str.strip().tolist()
        if s in xl.sheet_names else []
    )
    profs  = sorted(set(get_col('Profesores')) | set(df['Profesor'].unique()))
    cursos = sorted(set(get_col('Cursos'))    | set(df['Curso'].unique()))
    recs   = sorted(set(get_col('Recursos'))  |
                    set(df['Recurso'].str.split(', ').explode().dropna().unique()))
    return profs, cursos, recs

PROFESORES, CURSOS, RECURSOS = recalc_lists(df)

# ------------------------------------------------------------------
# 4) Funciones de guardado e informes
# ------------------------------------------------------------------
def atomic_save(df_save: pd.DataFrame) -> None:
    for _ in range(5):
        try:
            wb = load_workbook(EXCEL)
            if SHEET in wb.sheetnames:
                wb.remove(wb[SHEET])
            ws = wb.create_sheet(SHEET, 0)
            for c, col in enumerate(df_save.columns, 1):
                ws.cell(1, c, col)
            for r, row in enumerate(df_save.itertuples(index=False), 2):
                for c, val in enumerate(row, 1):
                    ws.cell(r, c, val)
            wb.save(EXCEL)
            break
        except PermissionError:
            time.sleep(0.2)
            pass

def build_report(df_report: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        # 1) Hoja de reservas
        df_report.to_excel(writer, sheet_name="Reservas", index=False)

        # 2) Hoja Uso recurso
        recurso = df_report["Recurso"].str.split(", ").explode()
        uso_rec = recurso.value_counts().rename("Reservas")
        uso_rec.to_excel(writer, sheet_name="Uso recurso")

        # 3) Hoja Uso profesor
        uso_prof = df_report["Profesor"].value_counts().rename("Reservas")
        uso_prof.to_excel(writer, sheet_name="Uso profesor")

        # 4) Hoja de gráficos
        workbook = writer.book
        chart_ws = workbook.add_worksheet("Gráficos")

        # Pie chart: Uso recurso
        chart1 = workbook.add_chart({"type": "pie"})
        n_rec   = uso_rec.size
        chart1.add_series({
            "name":       "Uso recurso",
            "categories": f"='Uso recurso'!$A$2:$A${n_rec+1}",
            "values":     f"='Uso recurso'!$B$2:$B${n_rec+1}",
            "data_labels": {"percentage": True},
        })
        chart1.set_title({"name": "Uso recurso"})
        chart_ws.insert_chart("A1", chart1)

        # Pie chart: Uso profesor
        chart2 = workbook.add_chart({"type": "pie"})
        n_prof  = uso_prof.size
        chart2.add_series({
            "name":       "Uso profesor",
            "categories": f"='Uso profesor'!$A$2:$A${n_prof+1}",
            "values":     f"='Uso profesor'!$B$2:$B${n_prof+1}",
            "data_labels": {"percentage": True},
        })
        chart2.set_title({"name": "Uso profesor"})
        chart_ws.insert_chart("H1", chart2)

    return buf.getvalue()

def build_ics(df_events: pd.DataFrame, prof: str) -> str:
    now = dt.datetime.utcnow()
    lines = [
        'BEGIN:VCALENDAR', 'VERSION:2.0',
        'PRODID:-//CAV//Horario Enlaces//ES', 'CALSCALE:GREGORIAN'
    ]
    for idx, row in df_events.iterrows():
        d = parse_date(row['Fecha'])
        hi = as_time(row['Hora inicio']); hf = as_time(row['Hora fin'])
        sd = dt.datetime.combine(d, hi); ed = dt.datetime.combine(d, hf)
        uid = f"{prof}-{idx}@cav.cl"
        lines += [
            'BEGIN:VEVENT',
            f'UID:{uid}',
            f'DTSTAMP:{now.strftime("%Y%m%dT%H%M%SZ")}',
            f'DTSTART;TZID=America/Santiago:{sd.strftime("%Y%m%dT%H%M%S")}',
            f'DTEND;TZID=America/Santiago:{ed.strftime("%Y%m%dT%H%M%S")}',
            f"SUMMARY:Reserva {row['Recurso']} ({row['Curso']})",
            f"DESCRIPTION:{row.get('Observaciones', '')}",
            'END:VEVENT'
        ]
    lines.append('END:VCALENDAR')
    return '\n'.join(lines)

# ------------------------------------------------------------------
# 5) Configuración de página y estilos
# ------------------------------------------------------------------
st.set_page_config('📅 HORARIO ENLACES CAV 💻', page_icon='', layout='wide')
CSS = f"""
html, body, [data-testid='stApp'] {{ background:#fff; color:#000; }}
.stButton>button, .stDownloadButton>button {{ background:{BURGUNDY}; color:#fff; }}
/* section[data-testid='stSidebar'] {{ display:none; }} */
[data-testid='stAppViewContainer']>div {{ padding:0 3rem; }}
label {{ color:#000!important; font-weight:600; }}
.stTableContainer .stTable th, .stTableContainer .stTable td {{
    border:1px solid #000!important; color:#000!important;
}}
[data-testid='stDataEditorContainer'] * {{ color:#000!important; }}
div[role='tablist'] > button[role='tab'] {{ color:#000!important; }}

@media (prefers-color-scheme: dark) {{
  html, body, [data-testid='stApp'] {{ background:#000; color:#fff; }}
  .stButton>button, .stDownloadButton>button {{ background:#5c0000; color:#fff; }}
  label {{ color:#fff!important; }}
  .stTableContainer .stTable th, .stTableContainer .stTable td {{
      border:1px solid #fff!important; color:#fff!important;
  }}
  [data-testid='stDataEditorContainer'] * {{ color:#fff!important; }}
  div[role='tablist'] > button[role='tab'] {{ color:#fff!important; }}
}}
"""

st.markdown(f"<style>{CSS}</style>", unsafe_allow_html=True)
#st.markdown(
#    f"<h1 style='text-align:center;font-size:2rem;color:{BURGUNDY};'>📅 HORARIO ENLACES CAV 💻</h1>",
#    unsafe_allow_html=True
#)

# ------------------------------------------------------------------
# 6) Función toast
# ------------------------------------------------------------------
def toast(msg: str, kind: str = 'info') -> None:
    icons = {
        'success': '✅',
        'error':   '❌',
        'warning': '⚠️',
        'info':    'ℹ️',
    }
    try:
        st.toast(msg, icon=icons.get(kind))
    except Exception:
        getattr(st, kind)(msg)

# ------------------------------------------------------------------
import streamlit.components.v1 as components
st.markdown(
    """
    <style>
      /* 1) Hacer que la Sidebar quede fija al hacer scroll */
      section[data-testid="stSidebar"] {
        position: sticky;
        top: 0;
        height: 100vh;
        overflow-y: auto;
      }

      /* 2) Ajustar ancho de la Sidebar si lo necesitas */
      section[data-testid="stSidebar"] {
        width: 260px !important;
      }

      /* 3) Aumentar el tamaño de letra del menú de navegación (radio buttons) */
      section[data-testid="stSidebar"] .stRadio label {
        font-size: 2rem !important;
        font-weight: 500 !important;
      }
    </style>
    """,
    unsafe_allow_html=True
)
# 3) Logo con ancho fijo (ajusta el valor a tu gusto)
st.sidebar.image("logo_CAV_2021-1.png", width=200)  # <- aquí cambias el tamaño
# 1) Datos de sesión
st.sidebar.markdown(f"**✅ Usuario:** {st.session_state.user}")
# 4) Título con tamaño de fuente personalizado
st.sidebar.markdown(
    "<h3 style='color:#800000; text-align:center; font-size:1.2rem; margin:0.5rem 0;'>"
    "💻HORARIO ENLACES CAV"
    "</h3>",
    unsafe_allow_html=True
)
# 5) Menú de navegación
role = st.session_state.role
if role == 'admin':
    pages = ['▶ Registrar','📂 Base datos','📅 Semana','🔧 Mantenimiento']
elif role == 'profesor':
    pages = ['▶ Registrar','📅 Semana']
else:
    pages = ['▶ Registrar']
page = st.sidebar.radio("📂 Navegar a:", pages, index=0)
# ——— Sidebar fijo con usuario, logout, logo y navegación ———



# 2) Botón Cerrar sesión
if st.sidebar.button("🚪  Cerrar sesión", use_container_width=True):
    for k in ["logged","user","role"]:
        st.session_state.pop(k, None)
    st.rerun()






# ————————————————————————————————————————————————

# ------------------------------------------------------------------
# 8) Sección ▶ Registrar (solo si page == '▶ Registrar')
# ------------------------------------------------------------------
if page == '▶ Registrar':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem;text-align:center'>⏩ Registrar nueva reserva</h2>",
        unsafe_allow_html=True
    )

    # — Carrusel + KPI — solo aquí —
    col_car, col_kpi = st.columns([2, 1], gap="small")

    with col_car:
        hoy   = dt.date.today()
        lunes = hoy - dt.timedelta(days=hoy.weekday())
        dias  = [lunes + dt.timedelta(days=i) for i in range(5)]
        nombres = ["Lunes","Martes","Miércoles","Jueves","Viernes"]
        dur     = 4

        slides = []
        for i, d in enumerate(dias):
            title = f"{nombres[i]} {d.strftime('%d/%m')}"
            sel   = df[df["Fecha"].apply(parse_date) == d]
            body  = (
                sel[["Hora inicio","Hora fin","Profesor","Curso","Recurso"]]
                .to_html(index=False)
                if not sel.empty else
                "<p>Sin reservas</p>"
            )
            slides.append(f"<div class='slide'><h5>{title}</h5>{body}</div>")

        html = f"""
        <style>
          .carousel {{ position: relative; width:100%; height:160px; overflow:hidden;
                        border:1px solid #ddd; border-radius:8px; }}
          .slide    {{ position:absolute; width:100%; height:100%; top:0; left:0;
                        opacity:0; animation: carousel {len(dias)*dur}s infinite; }}
          {"".join(f".slide:nth-child({i+1}){{animation-delay:{i*dur}s}} " for i in range(len(dias)))}
          @keyframes carousel {{0%{{opacity:0}} 5%{{opacity:1}} 20%{{opacity:1}}
                                25%{{opacity:0}} 100%{{opacity:0}}}}
          .slide h5    {{ margin:0.2rem; color:{BURGUNDY}; font-size:0.8rem;
                         text-align:center; }}
          .slide table {{ margin:0.2rem auto; width:95%; border-collapse:collapse;
                         font-size:0.6rem; text-align:center; }}
          .slide th,
          .slide td    {{ border:0.5px solid #000; padding:2px; }}
        </style>
        <div class="carousel">{''.join(slides)}</div>
        """
        components.html(html, height=160)

    with col_kpi:
        hoy       = dt.date.today()
        total     = len(df)
        hoy_count = (df["Fecha"].apply(parse_date) == hoy).sum()
        proxima   = df[df["Fecha"].apply(parse_date) >= hoy].iloc[0] if total else None

        st.markdown(
            f"<p style='font-size:1.3rem;margin:0'>{total}</p>"
            "<p style='font-size:0.8rem;color:gray;margin:0;'>🗓️ Total reservas</p>",
            unsafe_allow_html=True
        )
        st.markdown(
            f"<p style='font-size:1.1rem;margin:0'>{hoy_count}</p>"
            "<p style='font-size:0.8rem;color:gray;margin:0;'>📅 Reservas hoy</p>",
            unsafe_allow_html=True
        )
        valor = f"{proxima['Fecha']} {proxima['Hora inicio']}" if proxima is not None else "–"
        st.markdown(
            f"<p style='font-size:1.1rem;margin:0'>{valor}</p>"
            "<p style='font-size:0.8rem;color:gray;margin:0;'>⏰ Próxima reserva</p>",
            unsafe_allow_html=True
        )

    # — Formulario registrar —
    PROFESORES, CURSOS, RECURSOS = recalc_lists(df)
    def course_key(c: str) -> tuple[int,str]:
        cu = c.upper()
        if 'BÁSIC' in cu: return (0,cu)
        if 'MEDIO' in cu: return (1,cu)
        if 'DIF'   in cu: return (2,cu)
        return (3,cu)
    CURSOS = sorted(CURSOS, key=course_key)

    c1, c2 = st.columns(2, gap="small")
    with c1:
        fecha       = st.date_input('Fecha inicial', dt.date.today(), format='DD/MM/YYYY')
        recurrente  = st.checkbox("🔁 Hacer esta reserva recurrente")
        if recurrente:
            freq = st.selectbox("Frecuencia", ["Semanal","Diaria"])
            if freq == "Semanal":
                dias_sem = st.multiselect(
                    "Días de la semana",
                    nombres,
                    default=[nombres[fecha.weekday()]]
                )
            fecha_fin = st.date_input(
                "Repetir hasta",
                fecha + dt.timedelta(weeks=4),
                help="Fecha límite de la recurrencia"
            )
        hi = st.time_input('Hora inicio', BLOQUES[0][0])
        hf = st.time_input('Hora fin',    BLOQUES[0][1])

    with c2:
        prof   = st.selectbox('Profesor', PROFESORES)
        curso  = st.selectbox('Curso',    CURSOS)
        # bloqueo por mantenimiento igual que antes
        unavail = []
        if 'Mantenimientos' in xl.sheet_names:
            mant = xl.parse('Mantenimientos').fillna('')
            c0   = next((c for c in mant.columns if 'fecha' in c.lower() and 'inicio' in c.lower()), None)
            c1_  = next((c for c in mant.columns if 'fecha' in c.lower() and 'fin'    in c.lower()), None)
            if c0 and c1_:
                mant['Inicio'] = mant[c0].apply(parse_date)
                mant['Fin']    = mant[c1_].apply(parse_date)
                unavail = mant[
                    (mant['Inicio'] <= fecha) & (mant['Fin'] >= fecha)
                ]['Recurso'].astype(str).tolist()
        if unavail:
            st.warning(f"⚠️ Recursos en mantenimiento: {', '.join(unavail)}")
        available = [r for r in RECURSOS if r not in unavail]
        recs      = st.multiselect('Recursos', available)

    obs = st.text_area('Observaciones', height=68)

    if st.button('💾 Guardar reserva'):
        # — Validaciones básicas —
        if hi >= hf:
            toast('Hora inicio debe ser antes de fin.', 'warning'); st.stop()
        if not recs:
            toast('Selecciona al menos un recurso.', 'warning'); st.stop()

        # — Generamos la lista de fechas según recurrencia —
        fechas = []
        if recurrente:
            if freq == "Diaria":
                d = fecha
                while d <= fecha_fin:
                    fechas.append(d)
                    d += dt.timedelta(days=1)
            else:  # Semanal
                weekday_map = {n:i for i,n in enumerate(nombres)}
                sel_nums = [weekday_map[d] for d in dias_sem]
                d = fecha
                while d <= fecha_fin:
                    if d.weekday() in sel_nums:
                        fechas.append(d)
                    d += dt.timedelta(days=1)
        else:
            fechas = [fecha]

        # — Detectar choques (opcional: podrías comprobar cada fecha) —

        # — Construir DataFrame de nuevas reservas —
        rows = []
        for d in fechas:
            rows.append({
                "Fecha":         d.strftime(DATE_FMT),
                "Hora inicio":   hi.strftime("%H:%M"),
                "Hora fin":      hf.strftime("%H:%M"),
                "Profesor":      prof,
                "Curso":         curso,
                "Recurso":       ", ".join(recs),
                "Observaciones": obs.strip(),
            })
        new = pd.DataFrame(rows)

        # — Guardar todas las reservas de golpe —
        df = pd.concat([df, new], ignore_index=True)
        atomic_save(df)
        toast(f"✅ {len(new)} reservas creadas", "success")
 
# ------------------------------------------------------------------
# 9) Sección 📂 Base datos (solo admin)
# ------------------------------------------------------------------
elif page == '📂 Base datos' and role == 'admin':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem; text-align:center'>📂 Base datos de reservas</h2>",
        unsafe_allow_html=True
    )

    # 1) Prepara el DataFrame y la configuración de columnas
    editor = df.copy()
    editor['Fecha']       = pd.to_datetime(editor['Fecha'], dayfirst=True).dt.date
    editor['Hora inicio'] = editor['Hora inicio'].apply(as_time)
    editor['Hora fin']    = editor['Hora fin'].apply(as_time)

    cfg = {
        'Fecha':         st.column_config.DateColumn('Fecha', format='DD/MM/YYYY'),
        'Hora inicio':   st.column_config.TimeColumn('Hora inicio', format='HH:mm'),
        'Hora fin':      st.column_config.TimeColumn('Hora fin', format='HH:mm'),
        'Profesor':      st.column_config.SelectboxColumn('Profesor', options=PROFESORES),
        'Curso':         st.column_config.SelectboxColumn('Curso', options=CURSOS),
        'Recurso':       st.column_config.SelectboxColumn('Recurso', options=RECURSOS),
        'Observaciones': st.column_config.TextColumn('Observaciones'),
    }

    # 2) Editor de datos a todo ancho
    edited = st.data_editor(
        editor,
        hide_index=False,
        use_container_width=True,
        column_config=cfg,
        height=450
    )

    st.markdown("---")

    # 3) Primera fila: Guardar cambios y Eliminar registros
    action_col1, action_col2 = st.columns(2, gap="large")
    with action_col1:
        if st.button('💾 Guardar cambios', key='save_db_edits', use_container_width=True):
            atomic_save(edited)
            toast('✅ Cambios guardados.', 'success')
    with action_col2:
        to_drop = st.multiselect(
            'Seleccionar registros a eliminar',
            options=edited.index,
            key='drop_db'
        )
        if to_drop and st.button('🗑️ Eliminar registros', key='delete_db', use_container_width=True):
            new_df = edited.drop(index=to_drop).reset_index(drop=True)
            atomic_save(new_df)
            toast('✅ Registros eliminados.', 'success')

    st.markdown("---")

    # 4) Segunda fila: Informe Excel y Descargar calendario
    export_col1, export_col2 = st.columns(2, gap="large")
    with export_col1:
        rpt = build_report(df)
        st.download_button(
            label='📥 Informe Excel',
            data=rpt,
            file_name='informe_reservas.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True
        )
    with export_col2:
        prof_cal = st.selectbox('Seleccionar profesor (.ics)', PROFESORES, key='export_cal')
        ics_data = build_ics(df[df['Profesor']==prof_cal], prof_cal)
        st.download_button(
            label='📅 Descargar calendario',
            data=ics_data,
            file_name=f'{prof_cal}_reservas.ics',
            mime='text/calendar',
            use_container_width=True
        )

# ------------------------------------------------------------------
# 10) Sección 📅 Semana (admin y profesor)
# ------------------------------------------------------------------
elif page == '📅 Semana':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem;text-align:center'>📅 Vista semanal</h2>",
        unsafe_allow_html=True
    )
    # Selecciona cualquier fecha de la semana (cualquier día)
    fecha_ref = st.date_input(
        "Selecciona fecha de la semana",
        dt.date.today(),
        format="DD/MM/YYYY",
        help="Elige un día dentro de la semana que quieres visualizar"
    )
    # Calcular rango lunes-viernes
    start = fecha_ref - dt.timedelta(days=fecha_ref.weekday())  # lunes
    dias = [start + dt.timedelta(days=i) for i in range(5)]     # hasta viernes
    # Encabezados en español
    nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    cols = [f"{nombres[i]} {dias[i].strftime('%d/%m')}" for i in range(5)]
    rows = [f"Bloque {i+1}" for i in range(len(BLOQUES))]
    tabla = pd.DataFrame("", index=rows, columns=cols)
    # Rellenar datos en tabla
    for i, (hi, hf) in enumerate(BLOQUES):
        for j, d in enumerate(dias):
            sel = df[
                (df["Fecha"].apply(parse_date) == d)
                & df.apply(
                    lambda r: overlap(
                        hi, hf,
                        as_time(r["Hora inicio"]),
                        as_time(r["Hora fin"])
                    ),
                    axis=1,
                )
            ]
            if not sel.empty:
                lines = []
                for _, row in sel.iterrows():
                    lines.append(
                        f"{row['Hora inicio']}-{row['Hora fin']} | "
                        f"{row['Profesor']} | {row['Curso']} | "
                        f"{row['Recurso']} | {row['Observaciones']}"
                    )
                tabla.iat[i, j] = "\n\n".join(lines)
    # Aplicar estilo de colores
    import hashlib
    def get_color(text: str) -> str:
        h = hashlib.md5(text.encode("utf-8")).hexdigest()
        return f"#{h[:6]}"
    styled = (
        tabla.style
        .set_properties(**{"white-space": "pre-wrap"})
        .applymap(
    lambda v: f"background-color: {get_color(v)}; color:#f4f4f4 !important;"
    if isinstance(v, str) and v else ""
)
        .set_table_styles([
            {"selector": "th", "props": [("min-width", "150px")]},
            {"selector": "td", "props": [("min-width", "150px")]}  
        ])
        .set_table_attributes('style="table-layout:auto; width:100%;"')
    )
    st.markdown(styled.to_html(), unsafe_allow_html=True)

# ------------------------------------------------------------------
# 11) Sección 🔧 Mantenimiento (solo admin)
# ------------------------------------------------------------------
elif page == '🔧 Mantenimiento' and role == 'admin':
    st.markdown(
        "<h2 style='color:#000;font-size:1.5rem;text-align:center'>🔧 Gestión de Mantenimiento</h2>",
        unsafe_allow_html=True
    )
    def_tab = "Mantenimientos"
    if def_tab in xl.sheet_names:
        mant_df = xl.parse(def_tab).fillna("")
    else:
        mant_df = pd.DataFrame(
            columns=["Recurso", "FechaInicio", "HoraInicio", "FechaFin", "HoraFin"]
        )

    # --- Agregar nuevo mantenimiento ---
    st.subheader("Agregar nuevo mantenimiento")
    rsrc_maint = st.selectbox("Recurso", RECURSOS, key="maint_res")
    d1, d2 = st.columns(2)
    with d1:
        start_date = st.date_input(
            "Fecha inicio", dt.date.today(), key="mant_start_date"
        )
    with d2:
        end_date = st.date_input(
            "Fecha fin", dt.date.today(), key="mant_end_date"
        )
    t1, t2 = st.columns(2)
    with t1:
        start_time = st.time_input(
            "Hora inicio", BLOQUES[0][0], key="mant_start_time"
        )
    with t2:
        end_time = st.time_input(
            "Hora fin", BLOQUES[0][1], key="mant_end_time"
        )

    if st.button(
        "💾 Guardar mantenimiento", key="save_maint", use_container_width=True
    ):
        new_row = {
            "Recurso": rsrc_maint,
            "FechaInicio": start_date,
            "HoraInicio": start_time,
            "FechaFin": end_date,
            "HoraFin": end_time,
        }
        mant_df = pd.concat([mant_df, pd.DataFrame([new_row])], ignore_index=True)
        with pd.ExcelWriter(
            EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            mant_df.to_excel(writer, sheet_name=def_tab, index=False)
        st.success("✅ Mantenimiento registrado correctamente.")

    # --- Editar o eliminar mantenimientos ---
    st.subheader("Editar o eliminar mantenimientos")
    mant_editor = mant_df.copy()
    if "FechaInicio" in mant_editor.columns:
        mant_editor["FechaInicio"] = mant_editor["FechaInicio"].apply(parse_date)
    if "FechaFin" in mant_editor.columns:
        mant_editor["FechaFin"] = mant_editor["FechaFin"].apply(parse_date)
    if "HoraInicio" in mant_editor.columns:
        mant_editor["HoraInicio"] = mant_editor["HoraInicio"].apply(as_time)
    if "HoraFin" in mant_editor.columns:
        mant_editor["HoraFin"] = mant_editor["HoraFin"].apply(as_time)

    cfg_maint = {
        "Recurso": st.column_config.SelectboxColumn("Recurso", options=RECURSOS),
        "FechaInicio": st.column_config.DateColumn("Fecha inicio", format="DD/MM/YYYY"),
        "HoraInicio": st.column_config.TimeColumn("Hora inicio", format="HH:mm"),
        "FechaFin": st.column_config.DateColumn("Fecha fin", format="DD/MM/YYYY"),
        "HoraFin": st.column_config.TimeColumn("Hora fin", format="HH:mm"),
    }
    edited_maint = st.data_editor(
        mant_editor,
        hide_index=False,
        use_container_width=True,
        column_config=cfg_maint,
    )

    c3, c4 = st.columns(2)
    with c3:
        if st.button(
            "💾 Guardar cambios mantenimiento",
            key="save_maint_edits",
            use_container_width=True,
        ):
            with pd.ExcelWriter(
                EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                edited_maint.to_excel(writer, sheet_name=def_tab, index=False)
            st.success("✅ Cambios en mantenimiento guardados.")
    with c4:
        to_drop_maint = st.multiselect(
            "Seleccionar mantenimientos a eliminar",
            options=edited_maint.index,
            key="drop_maint",
        )
        if (
            st.button(
                "🗑️ Eliminar mantenimiento",
                key="delete_maint",
                use_container_width=True,
            )
            and to_drop_maint
        ):
            kept = edited_maint.drop(index=to_drop_maint).reset_index(drop=True)
            with pd.ExcelWriter(
                EXCEL, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                kept.to_excel(writer, sheet_name=def_tab, index=False)
            st.success("✅ Mantenimientos eliminados.")
            
           