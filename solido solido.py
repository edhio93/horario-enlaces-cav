# schedule_app.py – HORARIO ENLACES CAV (versión final completa)
# ---------------------------------------------------------------------------
# Ejecutar:
#     streamlit run schedule_app.py
# Requiere Streamlit ≥ 1.29
# ---------------------------------------------------------------------------
"""
Aplicación Streamlit para gestionar reservas de salas de enlaces y recursos
tecnológicos del Colegio Antonio Varas (2° semestre 2025).
Secciones:
▶ Registrar   – Añadir nueva reserva.
📂 Base datos  – Ver, editar, eliminar y descargar informe.
📅 Semana     – Vista semanal (lunes-viernes, 6 bloques).
"""
from __future__ import annotations
import datetime as dt
import time
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import xlsxwriter  # type: ignore

# Configuración global
EXCEL = "Recursos.xlsx"
SHEET = "Reservas"
DATE_FMT = "%d/%m/%Y"
BURGUNDY = "#800000"

# Bloques horarios (6 bloques)
BLOQUES = [
    (dt.time(8, 0),  dt.time(9, 30)),   # Bloque 1
    (dt.time(9, 45), dt.time(11, 15)),  # Bloque 2
    (dt.time(11, 30),dt.time(13, 0)),   # Bloque 3
    (dt.time(14, 0), dt.time(15, 30)),  # Bloque 4
    (dt.time(15, 45),dt.time(16, 30)),  # Bloque 5
    (dt.time(16, 30),dt.time(18, 30)),  # Bloque 6
]

# Helpers
def as_time(val: str | dt.time) -> dt.time:
    """
    Convierte un string o dt.time a dt.time.
    Acepta formatos 'HH:MM:SS', 'HH:MM' o 'H'.
    """
    if isinstance(val, dt.time):
        return val
    s = str(val).strip()
    for fmt in ("%H:%M:%S", "%H:%M", "%H"):
        try:
            return dt.datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Formato de hora inválido: {s}")

def overlap(hi1: dt.time, hf1: dt.time, hi2: dt.time, hf2: dt.time) -> bool:
    return max(hi1, hi2) < min(hf1, hf2)

# Inicializar archivo Excel si no existe
if not Path(EXCEL).exists():
    cols = ["Fecha", "Hora inicio", "Hora fin", "Profesor", "Curso", "Recurso", "Observaciones"]
    pd.DataFrame(columns=cols).to_excel(EXCEL, sheet_name=SHEET, index=False)

# Cargar datos
df = pd.read_excel(EXCEL, sheet_name=SHEET, dtype=str).fillna("")
xl = pd.ExcelFile(EXCEL)

# Dinámicas para formularios
def recalc_lists(df: pd.DataFrame) -> tuple[list[str], list[str], list[str]]:
    get_col = lambda s: xl.parse(s).iloc[:, 0].dropna().astype(str).str.strip().tolist() if s in xl.sheet_names else []
    profs = sorted(set(get_col("Profesores")) | set(df["Profesor"].unique()))
    cursos = sorted(set(get_col("Cursos")) | set(df["Curso"].unique()))
    recs = sorted(set(get_col("Recursos")) | set(df["Recurso"].str.split(", ").explode().dropna().unique()))
    return profs, cursos, recs

PROFESORES, CURSOS, RECURSOS = recalc_lists(df)

# Guardado atómico
def atomic_save(df_save: pd.DataFrame) -> None:
    for _ in range(5):
        try:
            wb = load_workbook(EXCEL)
            if SHEET in wb.sheetnames:
                wb.remove(wb[SHEET])
            ws = wb.create_sheet(SHEET, 0)
            for c, col in enumerate(df_save.columns, 1):
                ws.cell(1, c, col)
            for r, row in enumerate(df_save.itertuples(index=False), 2):
                for c, val in enumerate(row, 1):
                    ws.cell(r, c, val)
            wb.save(EXCEL)
            break
        except PermissionError:
            time.sleep(0.2)

# Informe Excel + gráficos
def build_report(df_report: pd.DataFrame) -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df_report.to_excel(writer, sheet_name="Reservas", index=False)
        recurso = df_report["Recurso"].str.split(", ").explode()
        recurso.value_counts().rename("Reservas").to_excel(writer, sheet_name="Uso recurso")
        df_report["Profesor"].value_counts().rename("Reservas").to_excel(writer, sheet_name="Uso profesor")
        wb = writer.book
        sh = wb.add_worksheet("Gráficos")
        # Gráfico uso recurso
        uc = recurso.value_counts()
        pie1 = wb.add_chart({"type": "pie"})
        pie1.add_series({
            "categories": f"='Uso recurso'!$A$2:$A${uc.size+1}",
            "values":     f"='Uso recurso'!$B$2:$B${uc.size+1}",
            "data_labels": {"percentage": True},
        })
        pie1.set_title({"name": "Uso recurso"})
        sh.insert_chart("A1", pie1)
        # Gráfico uso profesor
        up = df_report["Profesor"].value_counts()
        pie2 = wb.add_chart({"type": "pie"})
        pie2.add_series({
            "categories": f"='Uso profesor'!$A$2:$A${up.size+1}",
            "values":     f"='Uso profesor'!$B$2:$B${up.size+1}",
            "data_labels": {"percentage": True},
        })
        pie2.set_title({"name": "Uso profesor"})
        sh.insert_chart("H1", pie2)
    return buf.getvalue()

# Estilos y cabecera
st.set_page_config("HORARIO ENLACES CAV", page_icon="📅", layout="wide")
CSS = f"""
html, body, [data-testid='stApp'] {{ background:#fff; color:#000; }}
.stButton>button, .stDownloadButton>button {{ background:{BURGUNDY}; color:#fff; }}
section[data-testid='stSidebar'] {{ display:none; }}
[data-testid='stAppViewContainer']>div {{ padding:0 3rem; }}
label {{ color:#000!important; font-weight:600; }}
.stTableContainer .stTable th, .stTableContainer .stTable td {{ border:1px solid #000!important; color:#000!important; }}
[data-testid='stDataEditorContainer'] * {{ color:#000!important; }}
div[role='tablist'] > button[role='tab'] {{ color:#000!important; }}
/* Estilo para listas desplegables en data editor */
[data-testid='stDataEditorContainer'] select, [data-testid='stDataEditorContainer'] option {{ color:#fff !important; background:#000 !important; }}
"""
st.markdown(f"<style>{CSS}</style>", unsafe_allow_html=True)
st.markdown(f"<h1 style='text-align:center;color:{BURGUNDY};'>📅 HORARIO ENLACES CAV 💻</h1>", unsafe_allow_html=True)

# Pestañas
tab_reg, tab_db, tab_week = st.tabs(["▶ Registrar", "📂 Base datos", "📅 Semana"])

# Función toast
def toast(msg: str, kind: str = "info") -> None:
    icons = {"success": "✅", "error": "❌", "warning": "⚠️", "info": "ℹ️"}
    try:
        st.toast(msg, icon=icons.get(kind))
    except Exception:
        getattr(st, kind)(msg)

# ▶ Registrar
with tab_reg:
    st.markdown("<h2 style='color:#000;'>▶ Registrar nueva reserva</h2>", unsafe_allow_html=True)
    PROFESORES, CURSOS, RECURSOS = recalc_lists(df)
    # Ordenar cursos: Básico primero, luego Medio, luego Dif
    def course_key(c: str) -> tuple[int, str]:
        cu = c.upper()
        if "BÁSIC" in cu or "BASIC" in cu:
            return (0, cu)
        if "MEDIO" in cu:
            return (1, cu)
        if "DIF" in cu:
            return (2, cu)
        return (3, cu)
    CURSOS = sorted(CURSOS, key=course_key)
    c1, c2 = st.columns(2)
    with c1:
        fecha = st.date_input("Fecha", dt.date.today(), format="DD/MM/YYYY")
        hi = st.time_input("Hora inicio", BLOQUES[0][0])
        hf = st.time_input("Hora fin",    BLOQUES[0][1])
        obs = st.text_area("Observaciones", height=80)
    with c2:
        prof = st.selectbox("Profesor", PROFESORES)
        curso = st.selectbox("Curso", CURSOS)
        recs = st.multiselect("Recursos", RECURSOS)
    if st.button("💾 Guardar reserva", use_container_width=True):
        # Validaciones básicas
        if hi >= hf:
            toast("Hora inicio debe ser antes de fin.", "warning")
            st.stop()
        if not recs:
            toast("Selecciona al menos un recurso.", "warning")
            st.stop()
        # Formato de fecha para comparación
        fstr = fecha.strftime(DATE_FMT)
                # Verificar choques de recurso y horario por recurso
        for rsrc in recs:
            mask = (
                (df["Fecha"] == fstr)
                & df["Recurso"].str.contains(rsrc)
                & df.apply(
                    lambda r: overlap(
                        hi, hf,
                        as_time(r["Hora inicio"]),
                        as_time(r["Hora fin"])
                    ), axis=1
                )
            )
            if mask.any():
                toast(
                    f"Choque de horario y recurso: '{rsrc}' ya está reservado en este bloque.",
                    "error"
                )
                st.stop()
        # Crear nueva reserva
        new = pd.DataFrame([{
            "Fecha": fstr,
            "Hora inicio": hi.strftime("%H:%M"),
            "Hora fin": hf.strftime("%H:%M"),
            "Profesor": prof,
            "Curso": curso,
            "Recurso": ", ".join(recs),
            "Observaciones": obs.strip(),
        }])
        df = pd.concat([df, new], ignore_index=True)
        atomic_save(df)
        toast("✅ Reserva guardada.", "success")

# 📂 Base datos
with tab_db:
    st.markdown("<h2 style='color:#000;'>📂 Base datos de reservas</h2>", unsafe_allow_html=True)
    editor = df.copy()
    # Convertir columnas a tipos nativos para edición
    # Convertir Fecha (string o datetime) a date con formatos comunes
    def parse_date(val):
        if isinstance(val, dt.datetime):
            return val.date()
        if isinstance(val, dt.date):
            return val
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%Y %H:%M", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return dt.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Formato de fecha inválido: {s}")
    editor['Fecha'] = editor['Fecha'].apply(parse_date)
    editor['Hora inicio'] = editor['Hora inicio'].apply(as_time)
    editor['Hora fin'] = editor['Hora fin'].apply(as_time)
    # Config column_config
    cfg = {
        "Fecha": st.column_config.DateColumn("Fecha", format="DD/MM/YYYY"),
        "Hora inicio": st.column_config.TimeColumn("Hora inicio", format="HH:mm"),
        "Hora fin": st.column_config.TimeColumn("Hora fin", format="HH:mm"),
        "Profesor": st.column_config.SelectboxColumn("Profesor", options=PROFESORES),
        "Curso": st.column_config.SelectboxColumn("Curso", options=CURSOS),
    }
        # Configurar columna Recurso (lista desplegable si está disponible)
    if hasattr(st.column_config, "MultiSelectColumn"):
        cfg["Recurso"] = st.column_config.MultiSelectColumn("Recurso", options=RECURSOS)
    else:
        cfg["Recurso"] = st.column_config.TextColumn("Recurso")
    ed = st.data_editor(editor, hide_index=False, use_container_width=True, column_config=cfg)
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button("💾 Guardar cambios", use_container_width=True):
            df = ed
            atomic_save(df)
            PROFESORES, CURSOS, RECURSOS = recalc_lists(df)
            toast("Cambios guardados.", "success")
    with c2:
        drop = st.multiselect("Filas a eliminar", options=ed.index)
        if st.button("🗑️ Eliminar filas", use_container_width=True) and drop:
            df = ed.drop(index=drop).reset_index(drop=True)
            atomic_save(df)
            PROFESORES, CURSOS, RECURSOS = recalc_lists(df)
            toast("Fila(s) eliminadas.", "success")
    with c3:
        rpt = build_report(df)
        st.download_button(
            "📥 Informe Excel",
            rpt,
            file_name="informe_reservas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# 📅 Semana
with tab_week:
    st.markdown("<h2 style='color:#000;'>📅 Vista semanal</h2>", unsafe_allow_html=True)
    # Selecciona cualquier fecha de la semana (cualquier día)
    fecha_ref = st.date_input(
        "Selecciona fecha de la semana (cualquier día)",
        dt.date.today(),
        format="DD/MM/YYYY",
        help="Elige un día dentro de la semana que quieres visualizar"
    )
    # Calcular rango lunes-viernes
    start = fecha_ref - dt.timedelta(days=fecha_ref.weekday())  # lunes
    dias = [start + dt.timedelta(days=i) for i in range(5)]      # hasta viernes
    # Encabezados en español
    nombres = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes"]
    cols = [f"{nombres[i]} {dias[i].strftime('%d/%m')}" for i in range(5)]
    # Filas: bloques 1-6
    rows = [f"Bloque {i+1}" for i in range(len(BLOQUES))]
    tabla = pd.DataFrame("", index=rows, columns=cols)
    # Rellenar datos en tabla
    for i, (hi, hf) in enumerate(BLOQUES):
        for j, d in enumerate(dias):
            sel = df[
                (df['Fecha'].apply(parse_date) == d)
                & df.apply(
                    lambda r: overlap(
                        hi, hf,
                        as_time(r['Hora inicio']),
                        as_time(r['Hora fin'])
                    ),
                    axis=1
                )
            ]
            if not sel.empty:
                lines = []
                for _, row in sel.iterrows():
                    lines.append(
                        f"{row['Hora inicio']}-{row['Hora fin']} | {row['Profesor']} – {row['Curso']} | {row['Recurso']} | {row['Observaciones']}"
                    )
                tabla.iat[i, j] = "\n".join(lines)
        # Estilizar y mostrar tabla
    styled = (
        tabla.style
        .set_properties(**{'white-space': 'pre-wrap'})
        .set_table_styles([
            {'selector': 'th', 'props': [('min-width', '200px')]},
            {'selector': 'td', 'props': [('min-width', '200px')]}
        ])
        .set_table_attributes('style="table-layout:auto; width:100%;"')
    )
    # Mostrar tabla usando HTML para respetar estilos de ancho
    html = styled.to_html()
    st.markdown(html, unsafe_allow_html=True)
